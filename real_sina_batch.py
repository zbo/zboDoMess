from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
import csv
import requests
import time, os
import comlib


def get_request(bat_arr):
    query_string = ','.join(bat_arr)
    req_str = 'https://hq.sinajs.cn/list={0}'.format(query_string)
    return req_str


fileout = './content/in.xlsx'

if os.path.exists(fileout):
    os.remove(fileout)
    print('in file deleted')
else:
    print('no such file:%s' % fileout)

wb = Workbook()
ws = wb.active

import gen_finding_from_bao

protect_bottom = []
bao_history = {}
files = gen_finding_from_bao.get_all_files()
for f in files:
    data = gen_finding_from_bao.load_data(f)
    data.reverse()
    bao_history[data[0].name] = data

all_org_codes = comlib.get_orign_codes_from_xl()
all_sd_codes = comlib.get_sd_codes_from_xl()
all_codes = all_org_codes + all_sd_codes
batch_array = []
request_array = []
orign_code = []
for code in all_codes:
    orign_code.append(code)
    num = code.split('.')[0]
    mkt = code.split('.')[1]
    code = mkt.lower() + num
    batch_array.append(code)
    if len(batch_array) == 20:
        request_string = get_request(batch_array)
        request_array.append(request_string)
        batch_array = []
if len(batch_array) > 0:
    request_array.append(get_request(batch_array))

all_result = []
for request in request_array:
    print(request)
    time.sleep(2)
    header = {'Referer': 'https://finance.sina.com.cn'}
    response = requests.get(request, headers=header)
    ten_batch = response.text.split(';')
    for one in ten_batch:
        if len(one) > 100:
            all_result.append(one.split('=')[1])
        else:
            skip = 'skip:{0}'.format(one)

index = 0
for one in all_result:
    orign = orign_code[index]
    index = index + 1
    content_array = one.split(',')
    name = content_array[0][1:]
    open_price = content_array[1]
    yesterday_close_prise = content_array[2]
    now_price = content_array[3]
    high_price = content_array[4]
    low_price = content_array[5]
    buy_1 = content_array[10]
    sell_1 = content_array[20]
    gap = (float(now_price) - float(yesterday_close_prise)) * 100 / float(yesterday_close_prise)
    tmparr = orign.split('.')
    tmpcode = '{0}.{1}'.format(tmparr[1],tmparr[0]).lower()

    if tmpcode in bao_history:
        item = bao_history[tmpcode]
        if item[0].high<item[1].high and item[0].low<item[1].low and low_price>item[0].low:
            protect_bottom.append(orign)

    ws["A{0}".format(index)] = name
    ws["B{0}".format(index)] = orign
    ws["C{0}".format(index)] = gap
    thin = Side(border_style="thin", color="000000")
    ws["C{0}".format(index)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 3
    # fill in color
    # if buy_1 == '0':
    #     ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="000001")
    #     continue
    # if sell_1 == '0':
    #     ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="FF0000")
    #     continue
    if float(now_price) > float(yesterday_close_prise):
        ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="ec7c24")
        if gap > 9:
            ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="FF0000")
        elif gap > 5:
            ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="FF6600")
        elif gap < 1.5:
            ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="FFCC00")
        continue
    else:
        ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="6cac44")
        if gap < -9.5:
            ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="000001")
        elif gap < -5:
            ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="003300")
        elif gap > -1.5:
            ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="99CC00")
        continue

for hit in protect_bottom:
    print(hit)
wb.save(fileout)
