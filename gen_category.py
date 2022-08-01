#coding=utf-8
from cv2 import triangulatePoints
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side, Font
import csv, os
from openpyxl import Workbook
import gen_finding_from_store
import gen_finding_from_bao

filein = './content/大市场周期.xlsx'
fileout = './content/out.xlsx'
wb_out = Workbook()
wb = load_workbook(filename=filein)
high_sheet = wb['虎年高度']
track_sheet = wb['赛道']
category = []
existing = []
others = []
color = ['00FFFFFF','00FFCC99','00CCFFCC','00FFFF99','00CCCCFF',
         '00FF8080','00FF9900','00339966','00666699','00FF99CC',
         '00CC99FF','00993366','00FFFFCC','00CCFFFF','00660066','00FF8080']

if os.path.exists(fileout):
    os.remove(fileout)
    print('out file deleted')
else:
    print('no such file:%s' % fileout)


def get_categoty():
    index = 2
    category = set()
    while high_sheet['c{0}'.format(index)].value is not None:
        cate = high_sheet['c{0}'.format(index)].value
        category.add(cate)
        index = index + 1
    return category


def get_others():
    global existing
    with open('config/category.csv', 'r') as f:
        existing = []
        all_array = []
        reader = csv.reader(f)
        for row in reader:
            existing.append(row)
            all_array.extend(row)
    for s in category:
        if s in all_array:
            continue
        others.append(s)
    return others


def get_cate(cate):
    cates = []
    for item in existing:
        if cate in item:
            cates.append(item[0])
    if len(cates) >= 1:
        return cates
    else:
        return ['其他']


def copy_row(target_sheet, source_sheet, index):
    target_row = target_sheet.max_row + 1
    for c in range(1, source_sheet.max_column + 1):
        target_cell = target_sheet.cell(row=target_row, column=c)
        # print('index={0} column={1}'.format(index,c))
        source_cell = source_sheet.cell(row=index, column=c)
        target_cell.value = source_cell.value
        target_cell.fill = PatternFill("solid", fgColor=source_cell.fill.fgColor)
        thin = Side(border_style="thin", color="000000")
        target_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        target_cell.font = Font(name='华文楷体')
        if target_cell.column_letter not in ['A', 'B', 'C']:
            target_sheet.column_dimensions[target_cell.column_letter].width = 3


def copy_title(target_sheet, source_sheet):
    for c in range(1, source_sheet.max_column + 1):
        target_cell = target_sheet.cell(row=1, column=c)
        source_cell = source_sheet.cell(row=1, column=c)
        target_cell.value = source_cell.value
        # target_cell.fill = PatternFill("solid", fgColor=source_cell.fill.fgColor, bgColor=source_cell.fill.bgColor)
        target_cell.fill = PatternFill("solid", fgColor=source_cell.fill.fgColor)
        thin = Side(border_style="thin", color="000000")
        target_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def generate_cate_sheet():
    # global index, category, others, cate, name
    index = 2
    while high_sheet['c{0}'.format(index)].value is not None:
        cate = high_sheet['c{0}'.format(index)].value
        fix_cates = get_cate(cate)
        for fxc in fix_cates:
            copy_row(wb_out[fxc], high_sheet, index)
        index = index + 1
    for name in wb_out.sheetnames:
        copy_title(wb_out[name], high_sheet)
    sheet1 = wb_out['Sheet']
    wb_out.remove(sheet1)
    wb_out.save(fileout)


def gen_sheet():
    for item in existing:
        if '原始' not in wb_out.sheetnames:
            wb_out.create_sheet('原始')
        if '赛道' not in wb_out.sheetnames:
            wb_out.create_sheet('赛道')
        if '胜出' not in wb_out.sheetnames:
            wb_out.create_sheet('胜出')
        if '掉队' not in wb_out.sheetnames:
            wb_out.create_sheet('掉队')
        if '其他' not in wb_out.sheetnames:
            wb_out.create_sheet('其他')
        if '高度' not in wb_out.sheetnames:
            wb_out.create_sheet('高度')
        if '缩量' not in wb_out.sheetnames:
            wb_out.create_sheet('缩量')
        if item[0] not in wb_out.sheetnames:
            wb_out.create_sheet(title=item[0])
    wb_out.save(fileout)


def meet_high_condition(source_sheet, index):
    color = []
    for c in range(1, source_sheet.max_column + 1):
        source_cell = source_sheet.cell(row=index, column=c)
        color.append(source_cell.fill.fgColor)
    max_num = 0
    total_num = 0
    for color_item in color:
        if color_item.index == 'FFFF0000':
            total_num = total_num + 1
        else:
            if total_num > max_num:
                max_num = total_num
            total_num = 0
    return max_num >= 5


def generate_top_sheet():
    for i in range(1, high_sheet.max_row+1):
        if i == 1:
            copy_title(wb_out['高度'], high_sheet)
        else:
            if meet_high_condition(high_sheet, i):
                copy_row(wb_out['高度'], high_sheet, i)
    wb_out.save(fileout)


def generate_fx_sheet():
    fx_code_bao = gen_finding_from_bao.logic_fx()

def generate_sl_sheet():
    all_code_store = gen_finding_from_store.logic()
    all_code_bao = gen_finding_from_bao.logic()
    all_code = set(all_code_bao+all_code_store)
    for i in range(1, high_sheet.max_row+1):
        if i == 1:
            copy_title(wb_out['缩量'], high_sheet)
        elif high_sheet.cell(row=i, column=1).value is None:
            continue
        else:
            compare_code = high_sheet.cell(row=i, column=1).value.split('.')
            reverse_code = compare_code[1] + '.' + compare_code[0]
            if reverse_code in all_code:
                copy_row(wb_out['缩量'], high_sheet, i)
    wb_out.save(fileout)


def get_color_by_cate(sheet_instance, index):
    cate = sheet_instance.cell(row=index, column=3).value
    color_index = 0
    for i in range(len(existing)):
        if cate in existing[i]:
            color_index = i + 1
            break
    return color[color_index]

def fill_color(sheet_instance, index, color_str):
    sheet_instance.cell(row = index, column = 1).fill = PatternFill("solid", fgColor=color_str)
    sheet_instance.cell(row = index, column = 2).fill = PatternFill("solid", fgColor=color_str)
    sheet_instance.cell(row = index, column = 3).fill = PatternFill("solid", fgColor=color_str)

def generate_orign_sheet():
    for i in range(1, high_sheet.max_row+1):
        if high_sheet.cell(row=i,column=1).value is None:
            continue
        if i == 1:
            copy_title(wb_out['原始'], high_sheet)
        else:
            copy_row(wb_out['原始'], high_sheet, i)
            row_color = get_color_by_cate(high_sheet, i)
            fill_color(wb_out['原始'], i, row_color)
    wb_out.save(fileout)

def generate_split_sheet():
    out_code = []
    with open('out.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            out_code.append(row[0])
    for i in range(1, high_sheet.max_row+1):
        if high_sheet.cell(row=i,column=1).value is None:
            continue
        if i == 1:
            copy_title(wb_out['胜出'], high_sheet)
            copy_title(wb_out['掉队'], high_sheet)
        else:
            code_with_dot = high_sheet.cell(row=i,column=1).value
            code_arr = code_with_dot.split('.')
            code_reformat = code_arr[1]+code_arr[0]
            if code_reformat in out_code:
                copy_row(wb_out['掉队'], high_sheet, i)
            else:
                copy_row(wb_out['胜出'], high_sheet, i)
    wb_out.save(fileout)

def generate_tracking_sheet():
    target_sheet = wb_out['赛道']
    for i in range(1, track_sheet.max_row+1):
        if track_sheet.cell(row=i,column=1).value is None:
            continue
        if i == 1:
            copy_title(target_sheet, track_sheet)
        else:
            copy_row(target_sheet, track_sheet, i)
            row_color = get_color_by_cate(track_sheet, i)
            fill_color(target_sheet, i, row_color)
    wb_out.save(fileout)

def freeze_pan_for_all_sheet():
    for name in wb_out.sheetnames:
        wb_out[name].freeze_panes = 'D2'
    wb_out.save(fileout)


if __name__ == '__main__':
    category = get_categoty()
    # print(category)
    others = get_others()
    # print(others)
    gen_sheet()
    generate_cate_sheet()
    generate_orign_sheet()
    generate_top_sheet()
    generate_sl_sheet()
    generate_fx_sheet()
    generate_split_sheet()
    generate_tracking_sheet()
    freeze_pan_for_all_sheet()
