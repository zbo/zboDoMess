from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
import csv
import requests
import time,os


def get_request(bat_arr):
    query_string = ','.join(bat_arr)
    req_str = 'https://hq.sinajs.cn/list={0}'.format(query_string)
    return req_str

fileout = '/Users/zhubo/Documents/in.xlsx'

if os.path.exists(fileout):
    os.remove(fileout)
    print('in file deleted')
else:
    print('no such file:%s'%fileout)

wb = Workbook()
ws = wb.active

with open('list.csv', 'r') as f:
    batch_array = []
    request_array = []
    reader = csv.reader(f)
    orign_code = []
    for row in reader:
        if len(row) == 0:
            continue
        orign_code.append(row)
        code = row[0]
        num = code.split('.')[0]
        mkt = code.split('.')[1]
        code = mkt.lower() + num
        batch_array.append(code)
        if len(batch_array) == 20:
            request_string = get_request(batch_array)
            request_array.append(request_string)
            batch_array = []
    if len(batch_array) > 0:
        request_array.append(get_request(batch_array))

    all_result = []
    for request in request_array:
        print(request)
        time.sleep(1)
        header = {'Referer':'https://finance.sina.com.cn'}
        response = requests.get(request,headers=header)
        ten_batch = response.text.split(';')
        for one in ten_batch:
            if len(one) > 4:
                all_result.append(one.split('=')[1])

    index = 0
    for one in all_result:
        orign = orign_code[index]
        index = index + 1
        content_array = one.split(',')
        name = content_array[0][1:]
        open_price = content_array[1]
        yesterday_close_prise = content_array[2]
        now_price = content_array[3]
        high_price = content_array[4]
        low_price = content_array[5]
        buy_1 = content_array[10]
        sell_1 = content_array[20]
        gap = (float(now_price) - float(yesterday_close_prise)) * 100 / float(yesterday_close_prise)

        ws["A{0}".format(index)] = name
        ws["B{0}".format(index)] = orign[0]
        ws["C{0}".format(index)] = gap
        thin = Side(border_style="thin", color="000000")
        ws["C{0}".format(index)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 3
        ws.column_dimensions['C'].width = 3
        # fill in color
        # if buy_1 == '0':
        #     ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="000001")
        #     continue
        # if sell_1 == '0':
        #     ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="FF0000")
        #     continue
        if now_price > yesterday_close_prise:
            ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="ec7c24")
            if gap > 9:
                ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="FF0000")
            elif gap > 5:
                ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="FF6600")
            elif gap < 1.5:
                ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="FFCC00")
            continue
        if now_price <= yesterday_close_prise:
            ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="6cac44")
            if gap < -5:
                ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="003300")
            elif gap > -1.5:
                ws["C{0}".format(index)].fill = PatternFill("solid", fgColor="99CC00")
            continue

wb.save(fileout)
