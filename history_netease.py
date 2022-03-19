# coding=utf-8
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from io import StringIO
import os
import csv
import json
import time

import requests
from openpyxl.styles.builtins import styles

url_format = 'http://quotes.money.163.com/service/chddata.html?code={0}&start=20220101&end=20221231&fields=PCHG'
# url_format = 'http://quotes.money.163.com/service/chddata.html?code={0}&start=20201115&end=20210126&fields=PCHG'

wb = Workbook()
ws = wb.active


def gen_meta():
    with open('list.csv', 'r') as f:
        reader = csv.reader(f)
        all_codes = []
        all_urls = []
        for row in reader:
            all_codes.append(row)
            code = row[0].split('.')[0]
            mkt = '0'
            if row[0].split('.')[1].lower() == 'sz':
                mkt = '1'
            ncode = mkt + code
            url_formatted = url_format.format(ncode)
            all_urls.append(url_formatted)
        return all_urls, all_codes


class Stock:
    def __init__(self):
        self.each_day_change = []
        self.each_day = []
        self.name = ''
        self.code = ''

    def add_Item(self, item):
        self.each_day_change.append(item)

    def add_Day(self, day):
        self.each_day.append(day)


def fileExist(filename):
    if os.path.exists('./store/{0}'.format(filename)):
        return True
    return False


def fill_in(data_bag):
    lendays = len(data_bag[0].each_day_change)
    for data in data_bag:
        if len(data.each_day_change) < lendays:
            fill_size = lendays - len(data.each_day_change)
            array = ['FILL'] * fill_size
            data.each_day_change = data.each_day_change + array
    return data_bag


def gen_excel(data_bag):
    data_bag = fill_in(data_bag)
    lendays = len(data_bag[0].each_day_change)
    for i in range(lendays):
        day = data_bag[0].each_day[i]
        ws.cell(1, 2 + lendays - i, day[5:])
    cell_row_index = 1
    for data in data_bag:
        cell_row_index = cell_row_index + 1
        ws.cell(cell_row_index, 1, data.code)
        #
        ws.cell(cell_row_index, 2, data.name)
        thin = Side(border_style="thin", color="000000")
        for i in range(lendays):
            va = data.each_day_change[lendays - i - 1]
            filldata = str(va)
            if str(va) == "None":
                filldata = "停"
            if str(va) == "FILL":
                filldata = ""
            c = ws.cell(cell_row_index, 3 + i, filldata)
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if va == 'None' or va == 'FILL':
                va = 0
            va = float(va)
            if va < -9:
                c.fill = PatternFill("solid", fgColor="000001")
            elif va > 9:
                c.fill = PatternFill("solid", fgColor="FF0000")
            elif 5 < va <= 9:
                c.fill = PatternFill("solid", fgColor="FF6600")
            elif 1.5 < va <= 5:
                c.fill = PatternFill("solid", fgColor="ec7c24")
            elif 0 <= va < 1.5:
                c.fill = PatternFill("solid", fgColor="FFCC00")
            elif -1.5 <= va < 0:
                c.fill = PatternFill("solid", fgColor="99CC00")
            elif -5 <= va < -1.5:
                c.fill = PatternFill("solid", fgColor="6cac44")
            elif -9 <= va < -5:
                c.fill = PatternFill("solid", fgColor="003300")


if __name__ == '__main__':
    all_urls, all_codes = gen_meta()
    index = 0
    data_bag = []
    for u in all_urls:
        s = Stock()
        filename = all_codes[index][0] + '.csv'
        if not fileExist(filename):
            print(u)
            response = requests.get(u)
            f = StringIO(response.text)
            csv_reader = csv.reader(f, delimiter=',')
            f_local = open('./store/{0}'.format(filename), 'w')
            csv_writer = csv.writer(f_local)
            for line in csv_reader:
                csv_writer.writerow(line)
            f_local.close()
            time.sleep(1)

        f_local = open('./store/{0}'.format(filename), 'r')
        reader = csv.reader(f_local, delimiter=',')
        # print(filename)
        for row in reader:
            if reader.line_num == 1:
                continue
            s.add_Item(row[3])
            s.add_Day(row[0])
            s.name = row[2]
            s.code = all_codes[index][0]
        data_bag.append(s)
        f_local.close()
        index = index + 1
        print("processed {0}, current request grab {1} rows".format(
            index, len(s.each_day)))

    gen_excel(data_bag)
    wb.save('history.xlsx')
