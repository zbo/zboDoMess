from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side, Font
import csv,os,datetime
from openpyxl import Workbook

fileout = '/Users/zhubo/Documents/out.xlsx'
wb_out = load_workbook(filename=fileout)
filein = '/Users/zhubo/Documents/in.xlsx'
wb_in = load_workbook(filename=filein)


def get_date_num():
    now=datetime.datetime.now()
    return now.strftime('%d')

def get_cate_from_conf():
    all_cate=[]
    with open('config/category.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            all_cate.append(row[0])
    all_cate.append('其他')
    all_cate.append('高度')
    return all_cate

def find_cell_from_in(code):
    for row_num in range(1, wb_in['Sheet'].max_row+1):
        source_cell = wb_in['Sheet'].cell(row=row_num, column=2)
        source_code = source_cell.value
        if code == source_code:
            return wb_in['Sheet'].cell(row=row_num, column=3)

def update_sheet(tar_sheet, date_num):
    for tar_row_num in range(2,tar_sheet.max_row+1):
        tar_cell = tar_sheet.cell(row = tar_row_num, column = date_num)
        tar_code = tar_sheet.cell(row = tar_row_num, column = 1).value
        source_cell = find_cell_from_in(tar_code)
        tar_cell.value = source_cell.value
        tar_cell.fill = PatternFill("solid", fgColor=source_cell.fill.fgColor)
        thin = Side(border_style="thin", color="000000")
        tar_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        tar_cell.font = Font(name='华文楷体')

def update_out_wb(all_cate, date_num):
    for cate in all_cate:
        update_sheet(wb_out[cate], date_num)
    wb_out.save(fileout)

def find_date_num_from_out(date_num):
    max_col = wb_out['其他'].max_column
    for i in range(max_col,1,-1):
        if str(wb_out['其他'].cell(row=1, column=i).value) == date_num:
            col_char = wb_out['其他'].cell(row=1, column=i).column_letter
            col_num = i
            break
    print('find target column {0}'.format(col_char))
    return col_char, col_num

if __name__ == '__main__':
    all_cate = get_cate_from_conf()
    date_num = get_date_num()
    print(all_cate)
    print(date_num)
    col_char, col_num = find_date_num_from_out(date_num)
    update_out_wb(all_cate, col_num)
    wb_in.close()
    wb_out.close()
    