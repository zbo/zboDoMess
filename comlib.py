from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import csv

sl_scan_range = 10
filein = './content/大市场周期.xlsx'
highsheet = '虎年高度'

def get_codes():
    r = []
    with open('list.csv', 'r') as f:
        reader = csv.reader(f)
        index = 1
        for row in reader:
            code = row[0].split('.')
            r.append('{0}.{1}'.format(code[1], code[0]))
    return r


class Stock:
    def __init__(self):
        self.each_day_change = []
        self.each_day = []
        self.name = ''
        self.code = ''

    def add_Item(self, item):
        self.each_day_change.append(item)

    def add_Day(self, day):
        self.each_day.append(day)


def fill_stock(result):
    s = Stock()
    t = result.code[0].split('.')
    code = '{0}.{1}'.format(t[1],t[0])
    s.code = code
    s.name = 'BS_missing'
    s.each_day = result['date'].iloc[::-1].values
    result['close'] = result['close'].astype(float)
    result['preclose'] = result['preclose'].astype(float)
    result['change'] = (result['close']-result['preclose'])*100/result['preclose']
    result['change'] = result['change'].round(2)
    result['change'] = result['change'].astype(str).apply(lambda x:x.rstrip('0'))
    s.each_day_change = result['change'].iloc[::-1].values
    s.each_day_change = s.each_day_change.tolist()
    return s


def gen_excel(data_bag,ws):
    data_bag = fill_in(data_bag)
    lendays = len(data_bag[0].each_day_change)
    for i in range(lendays):
        day = data_bag[0].each_day[i]
        ws.cell(1, 2 + lendays - i, day[5:])
    cell_row_index = 1
    for data in data_bag:
        cell_row_index = cell_row_index + 1
        ws.cell(cell_row_index, 1, data.code)
        #
        ws.cell(cell_row_index, 2, data.name)
        thin = Side(border_style="thin", color="000000")
        for i in range(lendays):
            va = data.each_day_change[lendays - i - 1]
            filldata = str(va)
            if str(va) == "None":
                filldata = "停"
            if str(va) == "FILL":
                filldata = ""
            c = ws.cell(cell_row_index, 3 + i, filldata)
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if va == 'None' or va == 'FILL':
                va = 0
            va = float(va)
            if va < -9:
                c.fill = PatternFill("solid", fgColor="000001")
            elif va > 9:
                c.fill = PatternFill("solid", fgColor="FF0000")
            elif 5 < va <= 9:
                c.fill = PatternFill("solid", fgColor="FF6600")
            elif 1.5 < va <= 5:
                c.fill = PatternFill("solid", fgColor="ec7c24")
            elif 0 <= va <= 1.5:
                c.fill = PatternFill("solid", fgColor="FFCC00")
            elif -1.5 <= va < 0:
                c.fill = PatternFill("solid", fgColor="99CC00")
            elif -5 <= va < -1.5:
                c.fill = PatternFill("solid", fgColor="6cac44")
            elif -9 <= va < -5:
                c.fill = PatternFill("solid", fgColor="003300")
            if c.column_letter not in ['A', 'B']:
                ws.column_dimensions[c.column_letter].width = 3


def fill_in(data_bag):
    lendays = len(data_bag[0].each_day_change)
    for data in data_bag:
        if len(data.each_day_change) < lendays:
            fill_size = lendays - len(data.each_day_change)
            array = ['FILL'] * fill_size
            data.each_day_change = data.each_day_change + array
    return data_bag